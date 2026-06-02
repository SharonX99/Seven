"""
库存明细 API — GET /api/inventory, GET /api/stock-detail
"""

import calendar
import json
import os
from collections import defaultdict
from flask import Blueprint, jsonify, request, current_app
from backend.services.json_cache import JsonCache
from backend.routes.materials import _load_extra_materials, get_all_materials

inventory_bp = Blueprint('inventory', __name__)


def _load_warning_overrides(app):
    """加载用户自定义预警值"""
    path = app.config.get('WARNING_OVERRIDE_PATH', '')
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {}


def _save_warning_overrides(app, overrides):
    """保存用户自定义预警值"""
    path = app.config.get('WARNING_OVERRIDE_PATH', '')
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(overrides, f, ensure_ascii=False, indent=2)


@inventory_bp.route('/api/inventory', methods=['GET'])
def get_inventory():
    cache = JsonCache(current_app.config['JSON_CACHE_PATH'])
    data = cache.load()
    if not data:
        return jsonify({'error': '数据未加载'}), 500

    month = request.args.get('month', '')
    if not month:
        # 默认取最后一个月
        months = data.get('months', [])
        month = months[-1] if months else ''

    report = data.get('monthly_reports', {}).get(month, [])
    return jsonify({
        'month': month,
        'items': report,
        'total': len(report),
    })


@inventory_bp.route('/api/stock-detail', methods=['GET'])
def get_stock_detail():
    """
    按日期区间查询每日出入库明细
    GET /api/stock-detail?start_date=2026-05-01&end_date=2026-05-02
    """
    cache = JsonCache(current_app.config['JSON_CACHE_PATH'])
    data = cache.load()
    if not data:
        return jsonify({'error': '数据未加载'}), 500

    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    if not start_date or not end_date:
        return jsonify({'error': '请提供 start_date 和 end_date 参数'}), 400

    transactions = data.get('transactions', [])

    # 合并 entry_log 中的录入数据
    entry_log_path = current_app.config.get('ENTRY_LOG_PATH', '')
    entry_log = []
    if entry_log_path and os.path.exists(entry_log_path):
        try:
            with open(entry_log_path, 'r', encoding='utf-8') as f:
                el = json.load(f)
            entry_log = el.get('entries', [])
        except (json.JSONDecodeError, IOError):
            pass
    all_txns = list(transactions) + entry_log

    # 筛选日期区间内的交易
    filtered_txns = [t for t in all_txns if start_date <= t.get('date', '') <= end_date]

    # 按日期分组
    daily_groups = defaultdict(lambda: {'in': 0, 'out': 0, 'bad': 0, 'items': {}})

    for t in filtered_txns:
        d = t['date']
        if t['type'] == '入库':
            daily_groups[d]['in'] += t['quantity']
        elif t['type'] == '出库':
            daily_groups[d]['out'] += t['quantity']
        elif t['type'] == '不良':
            daily_groups[d]['bad'] += t['quantity']

        # 按物料汇总
        key = f"{t['material_code']}|{t['product_code']}"
        if key not in daily_groups[d]['items']:
            daily_groups[d]['items'][key] = {
                'material_code': t['material_code'],
                'product_code': t['product_code'],
                'in_qty': 0,
                'out_qty': 0,
                'bad_qty': 0,
            }
        if t['type'] == '入库':
            daily_groups[d]['items'][key]['in_qty'] += t['quantity']
        elif t['type'] == '出库':
            daily_groups[d]['items'][key]['out_qty'] += t['quantity']
        elif t['type'] == '不良':
            daily_groups[d]['items'][key]['bad_qty'] += t['quantity']

    # 构建返回数据
    daily_list = []
    total_in = 0
    total_out = 0
    total_bad = 0
    for d in sorted(daily_groups.keys()):
        g = daily_groups[d]
        total_in += g['in']
        total_out += g['out']
        total_bad += g['bad']
        daily_list.append({
            'date': d,
            'in_qty': g['in'],
            'out_qty': g['out'],
            'bad_qty': g['bad'],
            'net': g['in'] - g['out'] - g['bad'],
            'items': sorted(g['items'].values(), key=lambda x: x['material_code']),
        })

    return jsonify({
        'start_date': start_date,
        'end_date': end_date,
        'days': len(daily_list),
        'total_in': total_in,
        'total_out': total_out,
        'total_bad': total_bad,
        'total_net': total_in - total_out - total_bad,
        'daily': daily_list,
    })


@inventory_bp.route('/api/months', methods=['GET'])
def get_months():
    cache = JsonCache(current_app.config['JSON_CACHE_PATH'])
    data = cache.load()
    if not data:
        return jsonify({'error': '数据未加载'}), 500
    return jsonify(data.get('months', []))


@inventory_bp.route('/api/snapshot', methods=['GET'])
def get_snapshot():
    """
    按日期查看所有物料截至该日的库存快照
    GET /api/snapshot?date=2026-05-29
    """
    target_date = request.args.get('date', '')
    if not target_date:
        return jsonify({'error': '请提供 date 参数'}), 400
    return get_snapshot_data(target_date)


def get_snapshot_data(target_date):
    """获取指定日期的快照数据（供内部和导出共用）"""
    from flask import current_app, jsonify
    from backend.services.json_cache import JsonCache

    cache = JsonCache(current_app.config['JSON_CACHE_PATH'])
    data = cache.load()
    if not data:
        return jsonify({'error': '数据未加载'}), 500

    month_key = target_date[:7]
    if month_key not in data.get('monthly_reports', {}):
        return jsonify({'error': f'未找到 {month_key} 的报表数据'}), 404

    report = data.get('monthly_reports', {}).get(month_key, [])
    transactions = data.get('transactions', [])

    month_start = f'{month_key}-01'
    import calendar
    year = int(month_key[:4])
    month = int(month_key[5:7])
    last_day = calendar.monthrange(year, month)[1]
    month_end = f'{month_key}-{last_day:02d}'

    period_txns = [t for t in transactions if month_start <= t.get('date', '') <= target_date and t['date'] >= month_start]

    # 合并 entry_log 中的交易（用户录入但还未同步到 Excel 的）
    entry_log_path = current_app.config.get('ENTRY_LOG_PATH', '')
    if entry_log_path and os.path.exists(entry_log_path):
        try:
            with open(entry_log_path, 'r', encoding='utf-8') as f:
                el = json.load(f)
            for e in el.get('entries', []):
                if month_start <= e.get('date', '') <= target_date:
                    period_txns.append({
                        'date': e['date'],
                        'material_code': e['material_code'],
                        'product_code': e.get('product_code', ''),
                        'type': e['type'],
                        'quantity': e['quantity'],
                    })
        except (json.JSONDecodeError, IOError):
            pass

    period_in = defaultdict(int)
    period_out = defaultdict(int)
    period_bad = defaultdict(int)
    for t in period_txns:
        code = t['material_code']
        if t['type'] == '入库':
            period_in[code] += t['quantity']
        elif t['type'] == '出库':
            period_out[code] += t['quantity']
        elif t['type'] == '不良':
            period_bad[code] += t['quantity']

    snapshot_items = []
    for item in report:
        code = item['material_code']
        pi = period_in.get(code, 0)
        po = period_out.get(code, 0)
        pb = period_bad.get(code, 0)
        # 账面库存 = 期初 + 期间入库 - 期间出库
        cur_book = item['beginning_stock'] + pi - po
        # 实际库存 = 账面 - 报表不良 - 期间新增不良
        cur_bad = item['bad'] + pb
        cur_actual = cur_book - cur_bad

        snapshot_items.append({
            'material_code': code,
            'product_code': item['product_code'],
            'beginning_stock': item['beginning_stock'],
            'period_in': pi,
            'period_out': po,
            'period_bad': pb,
            'book_stock': cur_book,
            'bad': cur_bad,
            'actual_stock': cur_actual if cur_actual >= 0 else 0,
            'status': '库存充足' if cur_actual >= item.get('warning', 0) else '库存不足',
            'warning': item.get('warning', 0),
        })

    # 合并用户自定义预警值
    warning_overrides = _load_warning_overrides(current_app)
    for item in snapshot_items:
        code = item['material_code']
        if code in warning_overrides:
            item['warning'] = warning_overrides[code]

    # 计算缺口
    for item in snapshot_items:
        gap = item['warning'] - item['actual_stock']
        item['gap'] = max(0, gap)

    # 合并用户新增物料（在 Excel 中不存在的物料）
    extra_materials = _load_extra_materials(current_app)
    existing_codes = {item['material_code'] for item in snapshot_items}
    for code, mat in extra_materials.items():
        if code not in existing_codes:
            pi = period_in.get(code, 0)
            po = period_out.get(code, 0)
            pb = period_bad.get(code, 0)
            w = mat.get('warning', 0)
            if pi > 0 or po > 0 or pb > 0:
                cur_book = pi - po
                cur_actual = cur_book - pb
            else:
                cur_book = 0
                cur_actual = 0
            gap = max(0, w - cur_actual)
            snapshot_items.append({
                'material_code': code,
                'product_code': mat.get('product_code', ''),
                'beginning_stock': 0,
                'period_in': pi,
                'period_out': po,
                'period_bad': pb,
                'book_stock': cur_book,
                'bad': pb,
                'actual_stock': max(0, cur_actual),
                'status': '库存不足',
                'warning': w,
                'gap': gap,
            })

    snapshot_items.sort(key=lambda x: x['material_code'])

    total_beginning = sum(i['beginning_stock'] for i in snapshot_items)
    total_period_in = sum(i['period_in'] for i in snapshot_items)
    total_period_out = sum(i['period_out'] for i in snapshot_items)
    total_book = sum(i['book_stock'] for i in snapshot_items)
    total_actual = sum(i['actual_stock'] for i in snapshot_items)

    return jsonify({
        'date': target_date,
        'month': month_key,
        'items': snapshot_items,
        'total': len(snapshot_items),
        'summary': {
            'total_beginning': total_beginning,
            'total_period_in': total_period_in,
            'total_period_out': total_period_out,
            'total_book': total_book,
            'total_actual': total_actual,
        }
    })


@inventory_bp.route('/api/stock-detail/export', methods=['GET'])
def export_stock_detail_excel():
    """导出库存变动明细到 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    import tempfile

    # 复用 stock detail 数据
    resp = get_stock_detail()
    data = resp.get_json() if isinstance(resp, tuple) else resp.get_json()

    wb = Workbook()
    ws = wb.active
    ws.title = '库存变动明细'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['日期', '物料编码', '产品代码', '入库', '出库', '净变动']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row = 2
    for day in data.get('daily', []):
        for item in day.get('items', []):
            ws.cell(row, 1, day['date']).border = thin_border
            ws.cell(row, 2, item['material_code']).border = thin_border
            ws.cell(row, 3, item['product_code']).border = thin_border
            ws.cell(row, 4, item['in_qty']).border = thin_border
            ws.cell(row, 5, item['out_qty']).border = thin_border
            ws.cell(row, 6, item['in_qty'] - item['out_qty']).border = thin_border
            row += 1

    # 合计行
    ws.cell(row, 1, '合计').font = Font(bold=True)
    ws.cell(row, 4, data.get('total_in', 0)).font = Font(bold=True)
    ws.cell(row, 5, data.get('total_out', 0)).font = Font(bold=True)
    ws.cell(row, 6, data.get('total_net', 0)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'库存变动明细_{data.get("start_date","")}_{data.get("end_date","")}.xlsx'
    )


@inventory_bp.route('/api/snapshot/export', methods=['GET'])
def export_snapshot_excel():
    """导出指定日期库存快照到 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    target_date = request.args.get('date', '')
    if not target_date:
        return jsonify({'error': '请提供 date 参数'}), 400

    # 复用快照数据
    resp = get_snapshot_data(target_date)
    if isinstance(resp, tuple):
        return resp
    data = resp.get_json()

    wb = Workbook()
    ws = wb.active
    ws.title = f'库存快照_{target_date}'

    # 样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 表头
    headers = ['物料编码', '产品代码', '期初', '期间入库', '期间出库', '账面库存', '不良', '实际库存', '状态', '预警', '缺口']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据
    row = 2
    for item in data.get('items', []):
        ws.cell(row, 1, item['material_code']).border = thin_border
        ws.cell(row, 2, item['product_code']).border = thin_border
        ws.cell(row, 3, item['beginning_stock']).border = thin_border
        ws.cell(row, 4, item['period_in']).border = thin_border
        ws.cell(row, 5, item['period_out']).border = thin_border
        ws.cell(row, 6, item['book_stock']).border = thin_border
        ws.cell(row, 7, item['bad']).border = thin_border
        ws.cell(row, 8, item['actual_stock']).border = thin_border
        ws.cell(row, 9, item['status']).border = thin_border
        ws.cell(row, 10, item['warning']).border = thin_border
        ws.cell(row, 11, item['gap']).border = thin_border
        row += 1

    # 汇总行
    s = data.get('summary', {})
    summary_row = row
    ws.cell(summary_row, 1, '合计').font = Font(bold=True)
    ws.cell(summary_row, 3, s.get('total_beginning', 0)).font = Font(bold=True)
    ws.cell(summary_row, 4, s.get('total_period_in', 0)).font = Font(bold=True)
    ws.cell(summary_row, 5, s.get('total_period_out', 0)).font = Font(bold=True)
    ws.cell(summary_row, 6, s.get('total_book', 0)).font = Font(bold=True)
    ws.cell(summary_row, 8, s.get('total_actual', 0)).font = Font(bold=True)

    # 列宽
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10

    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    from flask import send_file
    return send_file(
        tmp.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'库存快照_{target_date}.xlsx'
    )


@inventory_bp.route('/api/gap-summary', methods=['GET'])
def get_gap_summary():
    """
    缺口汇总 — 查看所有有缺口的物料
    GET /api/gap-summary?date=2026-05-29
    """
    target_date = request.args.get('date', '')
    if not target_date:
        return jsonify({'error': '请提供 date 参数'}), 400

    resp = get_snapshot_data(target_date)
    if isinstance(resp, tuple):
        return resp
    data = resp.get_json()

    # 筛选有缺口的物料 (gap > 0)
    gap_items = [item for item in data.get('items', []) if item.get('gap', 0) > 0]
    total_gap = sum(item['gap'] for item in gap_items)

    return jsonify({
        'date': target_date,
        'items': gap_items,
        'count': len(gap_items),
        'total_gap': total_gap,
    })


@inventory_bp.route('/api/gap-summary/export', methods=['GET'])
def export_gap_summary_excel():
    """导出缺口汇总到 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    import tempfile

    target_date = request.args.get('date', '')
    if not target_date:
        return jsonify({'error': '请提供 date 参数'}), 400

    resp = get_gap_summary()
    data = resp.get_json() if isinstance(resp, tuple) else resp.get_json()

    wb = Workbook()
    ws = wb.active
    ws.title = f'缺口汇总_{target_date}'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['物料编码', '产品代码', '实际库存', '预警值', '缺口']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row = 2
    for item in data.get('items', []):
        ws.cell(row, 1, item['material_code']).border = thin_border
        ws.cell(row, 2, item['product_code']).border = thin_border
        ws.cell(row, 3, item['actual_stock']).border = thin_border
        ws.cell(row, 4, item['warning']).border = thin_border
        ws.cell(row, 5, item['gap']).border = thin_border
        ws.cell(row, 5).font = Font(bold=True, color='C62828')
        row += 1

    # 合计
    ws.cell(row, 1, '合计').font = Font(bold=True)
    ws.cell(row, 4, f'共 {data.get("count", 0)} 种').font = Font(bold=True)
    ws.cell(row, 5, data.get('total_gap', 0)).font = Font(bold=True, color='C62828')

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'缺口汇总_{target_date}.xlsx'
    )


@inventory_bp.route('/api/warning/update', methods=['POST'])
def update_warning():
    """
    更新物料预警值
    POST {"material_code": "1000087001", "warning": 50}
    """
    body = request.get_json(silent=True) or {}
    mat_code = body.get('material_code', '').strip()
    new_warning = body.get('warning')

    if not mat_code:
        return jsonify({'error': '物料编码不能为空'}), 400
    if new_warning is None or not isinstance(new_warning, (int, float)) or new_warning < 0:
        return jsonify({'error': '预警值必须是非负数字'}), 400

    overrides = _load_warning_overrides(current_app)
    overrides[mat_code] = int(new_warning)
    _save_warning_overrides(current_app, overrides)

    return jsonify({'success': True, 'material_code': mat_code, 'warning': int(new_warning)})
