"""
每日汇总 API — GET /api/daily
"""

import os
import json
from flask import Blueprint, jsonify, request, current_app
from backend.services.json_cache import JsonCache

daily_bp = Blueprint('daily', __name__)


def _load_entry_log(app):
    """加载用户录入日志"""
    path = app.config.get('ENTRY_LOG_PATH', '')
    if not path or not os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            el = json.load(f)
        return el.get('entries', [])
    except (json.JSONDecodeError, IOError):
        return []


@daily_bp.route('/api/daily', methods=['GET'])
def get_daily():
    cache = JsonCache(current_app.config['JSON_CACHE_PATH'])
    data = cache.load()
    if not data:
        return jsonify({'error': '数据未加载'}), 500

    date = request.args.get('date', '')
    if not date:
        return jsonify({'error': '请提供 date 参数'}), 400

    # 从 Excel 数据中获取该日交易
    transactions = data.get('transactions', [])

    # 合并 entry_log 中的录入数据
    entry_log = _load_entry_log(current_app)
    all_txns = list(transactions) + entry_log

    day_txns = [t for t in all_txns if t.get('date') == date]

    # 按物料汇总
    in_summary = {}
    out_summary = {}
    bad_summary = {}
    for t in day_txns:
        key = f"{t.get('material_code', '')}|{t.get('product_code', '')}"
        if t.get('type') == '入库':
            in_summary[key] = in_summary.get(key, 0) + t.get('quantity', 0)
        elif t.get('type') == '出库':
            out_summary[key] = out_summary.get(key, 0) + t.get('quantity', 0)
        elif t.get('type') == '不良':
            bad_summary[key] = bad_summary.get(key, 0) + t.get('quantity', 0)

    all_keys = set(list(in_summary.keys()) + list(out_summary.keys()) + list(bad_summary.keys()))
    summary = []
    for key in sorted(all_keys):
        code, prod = key.split('|', 1)
        in_q = in_summary.get(key, 0)
        out_q = out_summary.get(key, 0)
        bad_q = bad_summary.get(key, 0)
        summary.append({
            'material_code': code,
            'product_code': prod,
            'in_quantity': in_q,
            'out_quantity': out_q,
            'bad_quantity': bad_q,
            'net_change': in_q - out_q - bad_q,
        })

    return jsonify({
        'date': date,
        'summary': summary,
        'transactions': day_txns,
        'in_count': len(in_summary),
        'out_count': len(out_summary),
        'bad_count': len(bad_summary),
        'total_count': len(day_txns),
        'total_in_qty': sum(s['in_quantity'] for s in summary),
        'total_out_qty': sum(s['out_quantity'] for s in summary),
        'total_bad_qty': sum(s['bad_quantity'] for s in summary),
    })


@daily_bp.route('/api/daily/export', methods=['GET'])
def export_daily_excel():
    """导出当日出入库汇总到 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    import tempfile

    date = request.args.get('date', '')
    if not date:
        return jsonify({'error': '请提供 date 参数'}), 400

    # 复用 daily 数据
    resp = get_daily()
    data = resp.get_json() if isinstance(resp, tuple) else resp.get_json()

    wb = Workbook()
    ws = wb.active
    ws.title = f'日汇总_{date}'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['物料编码', '产品代码', '入库', '出库', '不良', '净变动']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row = 2
    for item in data.get('summary', []):
        ws.cell(row, 1, item['material_code']).border = thin_border
        ws.cell(row, 2, item['product_code']).border = thin_border
        ws.cell(row, 3, item['in_quantity']).border = thin_border
        ws.cell(row, 4, item['out_quantity']).border = thin_border
        ws.cell(row, 5, item.get('bad_quantity', 0)).border = thin_border
        ws.cell(row, 6, item['net_change']).border = thin_border
        row += 1

    # 合计行
    ws.cell(row, 1, '合计').font = Font(bold=True)
    ws.cell(row, 3, data.get('total_in_qty', 0)).font = Font(bold=True)
    ws.cell(row, 4, data.get('total_out_qty', 0)).font = Font(bold=True)
    ws.cell(row, 5, data.get('total_bad_qty', 0)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'日出入库汇总_{date}.xlsx'
    )
