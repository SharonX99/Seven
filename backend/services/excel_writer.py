"""
Excel 写入服务 — 将 entry_log 中未同步记录写入 Excel

策略：
1. 读取原始 Excel（非 data_only，保留公式）
2. 对未同步记录，找到对应 sheet、行、列
3. 更新每日出入库数值，同时更新汇总列
4. 保存 Excel
"""

import os
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """将录入数据写入 Excel 持久化"""

    # 静态列索引（同 ExcelReader）
    COL_MATERIAL_CODE = 1
    COL_PRODUCT_CODE = 2
    COL_BEGINNING = 3
    COL_TOTAL_IN = 4
    COL_TOTAL_OUT = 5
    COL_BOOK = 6
    COL_BAD = 7
    COL_ACTUAL = 8
    COL_STATUS = 9
    COL_WARNING = 10
    FIRST_DAILY_COL = 11

    def __init__(self, excel_path):
        self.excel_path = excel_path

    def find_sheet_and_cell(self, mat_code, entry_date):
        """
        根据物料编码和日期，找到对应 sheet 名称以及需要写入的列索引
        返回 (sheet_name, col_index) 或 (None, None)
        """
        target_month = entry_date.strftime('%Y-%m') if isinstance(entry_date, date) else entry_date[:7]
        target_day = entry_date.day if isinstance(entry_date, date) else int(entry_date.split('-')[2])

        wb = load_workbook(self.excel_path, data_only=True)

        # 找到匹配的 sheet
        sheet_name = self._find_sheet_for_month(wb, target_month)
        if not sheet_name:
            return None, None, None

        ws = wb[sheet_name]

        # 找到对应日期的列
        col_idx = self._find_date_column(ws, target_day)
        if not col_idx:
            return None, sheet_name, None

        # 找到对应物料行
        row_idx = self._find_material_row(ws, mat_code)
        if not row_idx:
            return None, sheet_name, None

        return sheet_name, col_idx, row_idx

    def write_entry(self, entry, sync_all=False):
        """
        写入一条或多条录入到 Excel
        entry: {"material_code":"...","date":"...","type":"入库/出库","quantity":N}
        返回写入结果 dict
        """
        from .excel_reader import ExcelReader

        wb = load_workbook(self.excel_path)
        entry_date = entry['date'] if isinstance(entry['date'], str) else entry['date'].strftime('%Y-%m-%d')
        target_month = entry_date[:7]
        target_day = int(entry_date.split('-')[2])
        mat_code = entry['material_code']
        entry_type = entry['type']  # '入库' or '出库'
        qty = entry['quantity']

        # 找 sheet
        excel_reader = ExcelReader(self.excel_path)
        sheet_names = excel_reader.get_sheet_names()
        sheet_name = self._find_sheet_for_month_str(sheet_names, target_month)
        if not sheet_name:
            return {'success': False, 'error': f'未找到 {target_month} 的 sheet'}

        ws = wb[sheet_name]

        # 找日期列
        col_pair = self._find_date_column_pair(ws, target_day)
        if not col_pair:
            return {'success': False, 'error': f'未找到 {entry_date} 的列'}

        in_col, out_col = col_pair

        # 找行
        row_idx = self._find_material_row(ws, mat_code)
        if not row_idx:
            return {'success': False, 'error': f'未找到物料 {mat_code}'}

        # 写入每日数据
        if entry_type == '入库':
            target_col = in_col
            total_col = self.COL_TOTAL_IN
        else:
            target_col = out_col
            total_col = self.COL_TOTAL_OUT

        old_daily = ws.cell(row=row_idx, column=target_col).value or 0
        new_daily = self._safe_int(old_daily) + qty
        ws.cell(row=row_idx, column=target_col, value=new_daily)

        # 更新汇总列（入库/出库总量）
        old_total = ws.cell(row=row_idx, column=total_col).value or 0
        new_total = self._safe_int(old_total) + qty
        ws.cell(row=row_idx, column=total_col, value=new_total)

        # 如果需要同步当前月份所有未同步记录
        if sync_all:
            self._sync_unsynced_to_sheet(ws, sheet_name, target_month, wb)

        wb.save(self.excel_path)
        wb.close()

        return {
            'success': True,
            'sheet': sheet_name,
            'row': row_idx,
            'col': target_col,
            'old_value': old_daily,
            'new_value': new_daily,
        }

    def batch_sync(self, entries):
        """
        批量同步未同步记录到 Excel
        entries: [entry_dict, ...]
        返回结果列表
        """
        results = []
        for entry in entries:
            result = self.write_entry(entry, sync_all=False)
            results.append({**result, 'entry_id': entry.get('id')})
        return results

    def _find_sheet_for_month(self, wb, target_month):
        """根据月份找 sheet"""
        from .excel_reader import ExcelReader
        reader = ExcelReader(self.excel_path)
        for sname in wb.sheetnames:
            month_key = reader._sheet_name_to_month(sname)
            if month_key == target_month:
                return sname
        return None

    def _find_sheet_for_month_str(self, sheet_names, target_month):
        """根据月份字符串找 sheet 名称"""
        from .excel_reader import ExcelReader
        reader = ExcelReader(self.excel_path)
        for sname in sheet_names:
            month_key = reader._sheet_name_to_month(sname)
            if month_key == target_month:
                return sname
        return None

    def _find_date_column(self, ws, target_day):
        """找到对应天的入库列（默认入库列）"""
        for col in range(self.FIRST_DAILY_COL, ws.max_column, 2):
            date_val = ws.cell(row=2, column=col).value
            if isinstance(date_val, datetime) and date_val.day == target_day:
                return col
        return None

    def _find_date_column_pair(self, ws, target_day):
        """
        找到对应天的列对 (in_col, out_col)
        Excel 中：入库和出库各占一列交替
        """
        for col in range(self.FIRST_DAILY_COL, ws.max_column, 2):
            date_val = ws.cell(row=2, column=col).value
            if isinstance(date_val, datetime) and date_val.day == target_day:
                return col, col + 1
        return None

    def _find_material_row(self, ws, mat_code):
        """找到物料对应的行号"""
        mat_code = str(mat_code).strip()
        for row in range(4, ws.max_row + 1):
            val = ws.cell(row=row, column=self.COL_MATERIAL_CODE).value
            if val and str(val).strip() == mat_code:
                return row
        return None

    def _safe_int(self, value):
        """安全转换整数"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return 0

    def _sync_unsynced_to_sheet(self, ws, sheet_name, target_month, wb):
        """（预留）同步所有未同步记录"""
        pass
