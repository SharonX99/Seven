"""
Excel 读取服务 — 解析苏州仓库库存管理 Excel 文件

数据布局（每个 sheet 代表一个月）：
- Row 1: 标题行
- Row 2: 列头（Col 1~10 静态 + Col 11+ 日期列）
- Row 3: 每日列进出标记（入库/出库交替）
- Row 4+: 数据行
  - Col 1: 物料编码
  - Col 2: 产品代码
  - Col 3: 期初库存
  - Col 4: 入库总量
  - Col 5: 出库总量
  - Col 6: 账面库存
  - Col 7: 仓库不良
  - Col 8: 实际库存
  - Col 9: 库存状态
  - Col 10: 预警标准
  - Col 11+: 每日数据（入库/出库交替列）
"""

import os
from datetime import datetime
from openpyxl import load_workbook


class ExcelReader:
    """Excel 读取器，解析库存管理表格"""

    # 静态列索引（1-based）
    COL_MATERIAL_CODE = 1
    COL_PRODUCT_CODE = 2
    COL_BEGINNING = 3
    COL_TOTAL_IN = 4
    COL_TOTAL_OUT = 5
    COL_BOOK = 6
    COL_BAD = 7
    COL_ACTUAL = 8
    COL_STATUS = 9
    COL_WARNING = 10
    FIRST_DAILY_COL = 11

    def __init__(self, excel_path):
        self.excel_path = excel_path
        self._wb = None

    def get_sheet_names(self):
        """获取所有 sheet 名称"""
        wb = self._load()
        return wb.sheetnames

    def parse_all(self):
        """
        解析所有 sheet，返回完整数据结构：
        {
            "materials": { "物料编码": {"material_code":"...", "product_code":"...", "warning": N} },
            "current_stock": [ {物料明细条目} ],
            "monthly_reports": { "2026-06": [条目] },
            "transactions": [ {"date":"...","material_code":"...","type":"入库/出库","quantity":N} ],
            "months": ["2025-08", ...],
            "last_updated": "2026-06-01T12:00:00"
        }
        """
        wb = self._load()
        materials = {}
        all_transactions = []
        monthly_reports = {}
        current_stock = []
        months = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            month_key = self._sheet_name_to_month(sheet_name)
            months.append(month_key)

            # 解析每日列头（R2日期 / R3进出标记）
            daily_columns = self._parse_daily_headers(ws)

            # 解析数据行
            report_items = []
            for row_idx in range(4, ws.max_row + 1):
                mat_code = ws.cell(row=row_idx, column=self.COL_MATERIAL_CODE).value
                if mat_code is None:
                    continue
                mat_code = str(mat_code).strip()

                prod_code = ws.cell(row=row_idx, column=self.COL_PRODUCT_CODE).value or ''
                prod_code = str(prod_code).strip()

                beginning = self._safe_int(ws.cell(row=row_idx, column=self.COL_BEGINNING).value)
                total_in = self._safe_int(ws.cell(row=row_idx, column=self.COL_TOTAL_IN).value)
                total_out = self._safe_int(ws.cell(row=row_idx, column=self.COL_TOTAL_OUT).value)
                book = self._safe_int(ws.cell(row=row_idx, column=self.COL_BOOK).value)
                bad = self._safe_int(ws.cell(row=row_idx, column=self.COL_BAD).value)
                actual = self._safe_int(ws.cell(row=row_idx, column=self.COL_ACTUAL).value)
                status_raw = ws.cell(row=row_idx, column=self.COL_STATUS).value or ''
                warning = self._safe_int(ws.cell(row=row_idx, column=self.COL_WARNING).value)

                # 解析每日出入库
                for date_str, col_idx, entry_type in daily_columns:
                    val = ws.cell(row=row_idx, column=col_idx).value
                    qty = self._safe_int(val)
                    if qty and qty > 0:
                        all_transactions.append({
                            'date': date_str,
                            'material_code': mat_code,
                            'product_code': prod_code,
                            'type': entry_type,
                            'quantity': qty,
                        })

                # 记录物料信息
                if mat_code not in materials:
                    materials[mat_code] = {
                        'material_code': mat_code,
                        'product_code': prod_code,
                        'warning': warning,
                    }

                report_items.append({
                    'material_code': mat_code,
                    'product_code': prod_code,
                    'beginning_stock': beginning,
                    'total_in': total_in,
                    'total_out': total_out,
                    'book_stock': book,
                    'bad': bad,
                    'actual_stock': actual,
                    'status': '库存充足' if '充足' in str(status_raw) else '库存不足',
                    'warning': warning,
                })

            monthly_reports[month_key] = report_items

            # 最后一个 sheet 作为当前库存
            if sheet_name == wb.sheetnames[-1]:
                current_stock = report_items

        months.sort()
        return {
            'materials': materials,
            'current_stock': current_stock,
            'monthly_reports': monthly_reports,
            'transactions': all_transactions,
            'months': months,
            'last_updated': datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
        }

    def _parse_daily_headers(self, ws):
        """
        解析每日列头，返回 [(date_str, col_idx, type), ...]
        - R2 行读取日期（每隔一列）
        - R3 行读取入库/出库标记
        Excel 中：col 11 = 入库, col 12 = 出库, col 13 = 入库, ...
        """
        daily_columns = []
        col = self.FIRST_DAILY_COL
        while col <= ws.max_column:
            date_cell = ws.cell(row=2, column=col).value
            if date_cell is None:
                break

            # 转换日期
            if isinstance(date_cell, datetime):
                date_str = date_cell.strftime('%Y-%m-%d')
            elif isinstance(date_cell, str):
                date_str = date_cell[:10].replace('/', '-')
            else:
                col += 1
                continue

            # 入库列
            in_type_raw = ws.cell(row=3, column=col).value
            out_type_raw = ws.cell(row=3, column=col + 1).value

            # 实际存放的是入库/出库标记
            in_type = '入库' if in_type_raw and '入' in str(in_type_raw) else None
            out_type = '出库' if out_type_raw and '出' in str(out_type_raw) else None

            if in_type:
                daily_columns.append((date_str, col, '入库'))
            if out_type:
                daily_columns.append((date_str, col + 1, '出库'))

            col += 2  # 每对占用2列

        return daily_columns

    def _sheet_name_to_month(self, sheet_name):
        """将 sheet 名称转为 YYYY-MM 格式"""
        # 去除多余空格
        name = sheet_name.strip()
        # 尝试提取年份和月份
        import re
        match = re.search(r'(\d{4})[年\s]*(\d+)[月\s]*', name)
        if match:
            return f'{match.group(1)}-{int(match.group(2)):02d}'
        return name

    def _safe_int(self, value):
        """安全转换整数"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return 0

    def _load(self):
        """懒加载工作簿"""
        if self._wb is None:
            self._wb = load_workbook(self.excel_path, data_only=True)
        return self._wb
